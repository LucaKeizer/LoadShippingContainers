# summary_exporter.py

# Standard Library Imports
from datetime import datetime
import json
import os
import re
import subprocess
import sys

# Third-party Imports
import openpyxl
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QMessageBox, QApplication, QFileDialog

# Local Application Imports
from src.utilities.utils import get_permanent_directory, open_folder


class SummaryExporter:
    """
    Handles exporting the loading summary to Excel (Loading Summary + Loading Order)
    and a JSON export (Istia Export) with relevant data for other systems.
    """

    def __init__(self, parent):
        """
        :param parent: Reference to the VisualizationPage (so we can access containers and packed items).
        """
        self.parent = parent  # VisualizationPage

    # ----------------------------
    # 1) Export Loading Summary (Excel)
    # ----------------------------

    def export_loading_summary(self):
        """Exports the loading summary, loading order, and synthetic pallets to an XLSX file."""
        wb, ws, loading_order_ws, synthetic_pallet_ws = self.setup_workbook()

        # Check for container data
        if not self.parent.containers_packed_items:
            QMessageBox.information(self.parent, "No Data", "No containers to export.")
            return

        # Build the summary sheet
        self.process_containers(wb, ws)
        # Build the loading order sheet
        self.process_loading_order_containers(loading_order_ws)
        # Build the synthetic pallets sheet
        self.process_synthetic_pallets(synthetic_pallet_ws)

        # Save and finalize
        self.finalize_workbook(wb)

    def setup_workbook(self):
        """
        Initializes the Excel workbook, sets column widths,
        and creates three sheets: 'Loading Summary', 'Loading Order', and 'Synthetic Pallets'.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Loading Summary"

        # Create a second sheet for loading order
        loading_order_ws = wb.create_sheet(title="Loading Order")
        # Create a third sheet for synthetic pallets
        synthetic_pallet_ws = wb.create_sheet(title="Synthetic Pallets")

        # Hide default gridlines
        for sheet in [ws, loading_order_ws, synthetic_pallet_ws]:
            sheet.sheet_view.showGridLines = False
            for col in range(1, 31):  # columns A to AD
                col_letter = get_column_letter(col)
                # Slightly narrower for specific columns if needed
                if col == 2:
                    sheet.column_dimensions[col_letter].width = 4
                else:
                    sheet.column_dimensions[col_letter].width = 11

        # Adjust specific column widths if necessary
        ws.column_dimensions['C'].width = 13
        loading_order_ws.column_dimensions['I'].width = 13  # Example adjustment

        # Insert titles
        self.insert_loading_summary_title(ws)
        self.insert_loading_order_title(loading_order_ws)
        self.insert_synthetic_pallets_title(synthetic_pallet_ws)

        return wb, ws, loading_order_ws, synthetic_pallet_ws

    def insert_loading_summary_title(self, ws):
        """Inserts 'Loading Summary' as a sheet title."""
        ws.merge_cells('D3:H3')
        title_cell = ws['D3']
        title_cell.value = "Loading Summary"
        title_cell.font = openpyxl.styles.Font(size=22)
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Add "cm//kg" notation in K3
        ws['K3'].value = "cm//kg"
        ws['K3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    def insert_loading_order_title(self, ws):
        """Inserts 'Loading Order' as a sheet title."""
        ws.merge_cells('D3:H3')
        title_cell = ws['D3']
        title_cell.value = "Loading Order"
        title_cell.font = openpyxl.styles.Font(size=22)
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Add "cm//kg" in I3
        ws['I3'].value = "cm//kg"
        ws['I3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    def insert_synthetic_pallets_title(self, ws):
        """Inserts 'Synthetic Pallets' as a sheet title."""
        ws.merge_cells('D3:H3')
        title_cell = ws['D3']
        title_cell.value = "Synthetic Pallets"
        title_cell.font = openpyxl.styles.Font(size=22)
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    def process_containers(self, wb, ws):
        """Populates the 'Loading Summary' sheet with container images, metrics, and items data."""
        current_row = 5
        container_type_counts = {}
        thin_border = self._create_thin_border()

        container_ids = sorted(self.parent.containers_packed_items.keys())
        for container_id in container_ids:
            self.process_single_container(ws, container_id, current_row, container_type_counts, thin_border)
            packed_items = self.parent.containers_packed_items[container_id]
            aggregated_items = self.aggregate_packed_items(packed_items)
            items_table_start_row = current_row + 11
            last_item_row = items_table_start_row + len(aggregated_items)
            current_row = last_item_row + 3

    def process_single_container(self, ws, container_id, current_row, container_type_counts, thin_border):
        """Inserts the container image, metrics, and item table for a single container."""
        self.parent.display_container(container_id)
        QApplication.processEvents()

        # Retrieve container and set camera distance based on container length.
        container = self.parent.containers.get(container_id)
        if not container:
            return

        # Use a wider camera angle (greater distance) for longer containers (e.g., 40ft).
        # Adjust the threshold as appropriate. Here, if container.length >= 1000 cm, we assume it's a 40ft container.
        distance = 800
        if container.length >= 1000:
            distance = 1300

        # Position camera
        self.parent.right_panel.view.setCameraPosition(elevation=45, azimuth=45, distance=distance)
        QApplication.processEvents()

        # Update walls, then capture
        self.parent.right_panel.update_wall_visibility()
        QApplication.processEvents()

        # Grab container image
        img = self.parent.right_panel.view.grabFramebuffer()
        export_dir = get_permanent_directory("Export Summaries\\Container Images")
        os.makedirs(export_dir, exist_ok=True)
        image_filename = f"container_{container_id}_diagonal1.png"
        image_path = os.path.join(export_dir, image_filename)
        img.save(image_path, "PNG")

        container = self.parent.containers.get(container_id)
        if not container:
            return

        container_type = container.container_type
        # Track count per container type (for labeling)
        container_type_counts[container_type] = container_type_counts.get(container_type, 0) + 1
        label = f"{container_type_counts[container_type]}({container_type})"

        # Insert container label
        label_cell = ws.cell(row=current_row - 1, column=3, value=label)
        label_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Insert container image
        img_excel = OpenPyXLImage(image_path)
        img_excel.height = 200
        img_excel.width = 300
        ws.add_image(img_excel, f"C{current_row}")

        # Insert metrics & items
        packed_items = self.parent.containers_packed_items[container_id]
        aggregated_items = self.aggregate_packed_items(packed_items)
        metrics = self.calculate_metrics(container, packed_items, aggregated_items)
        self.insert_metrics(ws, metrics, current_row, start_col=8)
        self.insert_items_table(ws, aggregated_items, current_row + 11)

        # Apply border to the relevant area
        items_table_start_row = current_row + 11
        last_item_row = items_table_start_row + len(aggregated_items)
        for row in ws.iter_rows(min_row=4, max_row=last_item_row, min_col=2, max_col=10):
            for cell in row:
                cell.border = thin_border

    def _create_thin_border(self):
        """Helper to create a thin gray border style."""
        light_gray = "CCCCCC"
        thin_side = openpyxl.styles.Side(border_style="thin", color=light_gray)
        return openpyxl.styles.Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def process_loading_order_containers(self, loading_order_ws):
        """Populates the 'Loading Order' sheet with images and item placement order."""
        thin_border = self._create_thin_border()
        container_ids = sorted(self.parent.containers_packed_items.keys())
        if not container_ids:
            return

        for idx, container_id in enumerate(container_ids, start=1):
            start_col = 3 + (idx - 1) * 7  # Each container uses 7 columns
            self.process_single_loading_order_container(loading_order_ws, container_id, start_col, thin_border)

    def process_single_loading_order_container(self, ws, container_id, start_col, thin_border):
        """Inserts a container image and the loading order table for that container."""
        self.parent.display_container(container_id)
        QApplication.processEvents()

        # Retrieve container and set camera distance based on container length.
        container = self.parent.containers.get(container_id)
        if not container:
            return

        # Determine distance based on container length (wider angle for longer containers)
        distance = 800
        if container.length >= 1000:
            distance = 1200

        self.parent.right_panel.view.setCameraPosition(elevation=45, azimuth=45, distance=distance)
        QApplication.processEvents()

        self.parent.right_panel.update_wall_visibility()
        QApplication.processEvents()

        img = self.parent.right_panel.view.grabFramebuffer()
        export_dir = get_permanent_directory("Export Summaries\\Container Images")
        os.makedirs(export_dir, exist_ok=True)
        image_filename = f"container_{container_id}_diagonal1.png"
        image_path = os.path.join(export_dir, image_filename)
        img.save(image_path, "PNG")

        container = self.parent.containers.get(container_id)
        if not container:
            return

        # Insert container label above the image
        label_row = 4
        image_row = 5
        table_start_row = image_row + 11
        label_cell = ws.cell(row=label_row, column=start_col, value=f"1({container.container_type})")
        label_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Insert image
        img_excel = OpenPyXLImage(image_path)
        img_excel.height = 200
        img_excel.width = 300
        ws.add_image(img_excel, f"{get_column_letter(start_col)}{image_row}")

        # Adjust columns
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 13
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 14
        ws.column_dimensions[get_column_letter(start_col + 3)].width = 14
        ws.column_dimensions[get_column_letter(start_col + 4)].width = 12

        # Insert loading order table by sorting items by (x, y, z)
        packed_items = self.parent.containers_packed_items[container_id]
        sorted_items = sorted(packed_items, key=lambda pi: (pi.position[0], pi.position[1], pi.position[2]))
        headers = ["Order", "Product Code", "Position", "Dimensions", "Weight (kg)"]
        for i, header in enumerate(headers):
            hdr_cell = ws.cell(row=table_start_row, column=start_col + i, value=header)
            hdr_cell.font = openpyxl.styles.Font(bold=True)
            hdr_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        data_start = table_start_row + 1
        for idx_item, p_item in enumerate(sorted_items, start=1):
            r = data_start + idx_item - 1
            order_cell = ws.cell(row=r, column=start_col, value=idx_item)
            sku_cell = ws.cell(row=r, column=start_col + 1, value=p_item.sku)
            pos_text = f"({int(p_item.position[0])}, {int(p_item.position[1])}, {int(p_item.position[2])})"
            pos_cell = ws.cell(row=r, column=start_col + 2, value=pos_text)
            dim_text = f"{int(p_item.size[0])} x {int(p_item.size[1])} x {int(p_item.size[2])}"
            dim_cell = ws.cell(row=r, column=start_col + 3, value=dim_text)
            w_cell = ws.cell(row=r, column=start_col + 4, value=f"{p_item.weight:.2f}")

            for c in [order_cell, sku_cell, pos_cell, dim_cell, w_cell]:
                c.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        last_row = table_start_row + len(sorted_items)
        for row in ws.iter_rows(min_row=4, max_row=last_row, min_col=start_col - 1, max_col=start_col + 6):
            for cell in row:
                cell.border = thin_border

    def aggregate_packed_items(self, packed_items):
        """Returns a dict of aggregated data by SKU with original quantity and carton information."""
        aggregated = {}
        # Get reference to parent_items for original info lookup
        parent_items = self.parent.items
        
        # Count occurrences of each SKU in packed_items
        sku_counts = {}
        for pi in packed_items:
            sku = pi.sku
            if sku in sku_counts:
                sku_counts[sku] += 1
            else:
                sku_counts[sku] = 1
        
        # Build lookup dictionary for original item info
        original_items_dict = {}
        for item in parent_items:
            base_sku = self.parent.get_base_sku(item.sku)
            if base_sku in original_items_dict:
                original_items_dict[base_sku]['original_quantity'] += item.quantity
                original_items_dict[base_sku]['cartons'] += getattr(item, 'cartons', 0)
            else:
                original_items_dict[base_sku] = {
                    'original_quantity': item.quantity,
                    'cartons': getattr(item, 'cartons', 0),
                    'is_carton_item': getattr(item, 'is_carton_item', False)
                }
        
        # Aggregate packed items
        for pi in packed_items:
            sku = pi.sku
            base_sku = self.parent.get_base_sku(sku)
            
            # Get original info
            original_info = original_items_dict.get(base_sku, {})
            
            if sku not in aggregated:
                # Use the count of occurrences as the quantity
                display_quantity = sku_counts.get(sku, 1)
                display_cartons = original_info.get('cartons', 0)
                
                aggregated[sku] = {
                    'quantity': display_quantity,
                    'cartons': display_cartons,
                    'length': pi.size[0],
                    'width': pi.size[1],
                    'height': pi.size[2],
                    'weight': pi.weight
                }
            else:
                aggregated[sku]['weight'] += pi.weight
        return aggregated
        
    def process_synthetic_pallets(self, ws):
        """
        Populates the 'Synthetic Pallets' sheet with combined, europallet, and mixed pallets and their contained items.
        Data is inserted starting from column C, leaving columns A and B empty.

        This version applies an alternating light blue background (columns C to J) for each distinct Pallet SKU.
        """
        from openpyxl.styles import PatternFill

        # Define a helper to recursively flatten all contained sub-items
        def _flatten_contained_items(parent_item):
            results = []
            for child in parent_item.contained_items:
                results.append(child)
                if child.contained_items:
                    results.extend(_flatten_contained_items(child))
            return results

        # Define headers starting at row 4, column C (3)
        headers = [
            "Pallet SKU",
            "Pallet Type",
            "Pallet Weight (kg)",
            "Contained Item SKU",
            "Contained Item Quantity",
            "Contained Item Dimensions (LxWxH cm)",
            "Contained Item Weight (kg)"
        ]
        for col_idx, header in enumerate(headers, start=3):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        current_row = 5  # Start writing data from row 5

        # Prepare alternating fill variables
        last_sku = None
        fill_flag = False
        fill_blue = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")

        # Access data_manager to retrieve combined, europallet, and mixed pallets
        data_manager = self.parent.parent.data_manager
        combined_items = data_manager.combined_pallets
        mixed_items = [it for it in data_manager.items if it.sku.startswith("MIXED-")]
        euro_items = [it for it in data_manager.items if it.sku.startswith("EuroP-")]
        all_synthetic_pallets = list(combined_items) + mixed_items + euro_items

        for pallet_item in all_synthetic_pallets:
            # Toggle fill when the Pallet SKU changes
            if pallet_item.sku != last_sku:
                fill_flag = not fill_flag
                last_sku = pallet_item.sku

            # Determine pallet type based on SKU
            if pallet_item.sku.startswith("CombP-"):
                pallet_type = "Combined"
            elif pallet_item.sku.startswith("EuroP-"):
                pallet_type = "Combined Europallet"
            elif pallet_item.sku.startswith("MIXED-"):
                pallet_type = "Mixed"
            else:
                continue

            # Flatten contained items
            all_subitems = _flatten_contained_items(pallet_item)
            if not all_subitems:
                continue

            for child in all_subitems:
                ws.cell(row=current_row, column=3, value=pallet_item.sku)         # Pallet SKU
                ws.cell(row=current_row, column=4, value=pallet_type)             # Pallet Type
                ws.cell(row=current_row, column=5, value=round(pallet_item.weight, 2))      # Pallet Weight rounded to 2 decimals

                ws.cell(row=current_row, column=6, value=child.sku)                # Contained Item SKU
                ws.cell(row=current_row, column=7, value=child.quantity)           # Contained Item Quantity
                dimensions = f"{int(child.length)}x{int(child.width)}x{int(child.height)}"
                ws.cell(row=current_row, column=8, value=dimensions)               # Dimensions
                ws.cell(row=current_row, column=9, value=child.weight)             # Contained Item Weight

                # For columns C (3) through J (10), set alignment and (if toggled) fill
                for col in range(3, 10):
                    cell = ws.cell(row=current_row, column=col)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    if fill_flag:
                        cell.fill = fill_blue
                    else:
                        cell.fill = openpyxl.styles.PatternFill()  # Reset fill

                current_row += 1

        # Adjust column widths for better readability
        from openpyxl.utils import get_column_letter
        for col_idx in range(3, 10):
            col_letter = get_column_letter(col_idx)
            if col_letter == 'C':
                ws.column_dimensions[col_letter].width = 15
            elif col_letter == 'D':
                ws.column_dimensions[col_letter].width = 20
            elif col_letter == 'E':
                ws.column_dimensions[col_letter].width = 18
            elif col_letter == 'F':
                ws.column_dimensions[col_letter].width = 20
            elif col_letter == 'G':
                ws.column_dimensions[col_letter].width = 23
            elif col_letter == 'H':
                ws.column_dimensions[col_letter].width = 37
            elif col_letter == 'I':
                ws.column_dimensions[col_letter].width = 26

        # Apply borders to all cells with data
        thin_border = self._create_thin_border()
        for row in ws.iter_rows(min_row=4, max_row=current_row - 1, min_col=3, max_col=9):
            for cell in row:
                cell.border = thin_border

    def calculate_metrics(self, container, packed_items, aggregated):
        """Returns a list of (label, value) for container metrics."""
        return [
            ("Volume Use Rate (%)", self.calculate_volume_use_rate(container, packed_items)),
            ("Weight Use Rate (%)", self.calculate_weight_use_rate(container, packed_items)),
            ("Goods Quantity", self.calculate_goods_quantity(aggregated)),
            ("Goods Volume (m³)", self.calculate_goods_volume(aggregated)),
            ("Goods Weight (kg)", self.calculate_goods_weight(aggregated)),
            ("Remainder Lengthwise", self.calculate_remainder_lengthwise(container, packed_items)),
            ("Remainder Widthwise", self.calculate_remainder_widthwise(container, packed_items)),
            ("Remainder Heightwise", self.calculate_remainder_heightwise(container, packed_items)),
        ]

    def insert_metrics(self, ws, metrics, start_row, start_col):
        """Inserts container metrics into the summary sheet."""
        row = start_row
        for label, val in metrics:
            lbl_cell = ws.cell(row=row, column=start_col, value=label)
            val_cell = ws.cell(row=row, column=start_col + 2, value=val)
            lbl_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            val_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            row += 1

    def insert_items_table(self, ws, aggregated_items, start_row):
        """Inserts an items table for each container with an added Cartons column."""
        headers = ["Product Code", "Quantity", "Cartons", "Length", "Width", "Height", "Weight"]
        for col_idx, hdr in enumerate(headers, start=3):
            hdr_cell = ws.cell(row=start_row, column=col_idx, value=hdr)
            hdr_cell.font = openpyxl.styles.Font(bold=True)
            hdr_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        current_row = start_row + 1
        for sku, data in aggregated_items.items():
            ws.cell(row=current_row, column=3, value=sku)
            ws.cell(row=current_row, column=4, value=data['quantity'])
            ws.cell(row=current_row, column=5, value=data['cartons'])
            ws.cell(row=current_row, column=6, value=data['length'])
            ws.cell(row=current_row, column=7, value=data['width'])
            ws.cell(row=current_row, column=8, value=data['height'])
            ws.cell(row=current_row, column=9, value=data['weight'])

            for col in range(3, 10):  # Updated to include new column
                cell = ws.cell(row=current_row, column=col)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            current_row += 1

    def finalize_workbook(self, wb):
        """Saves the workbook using the loading_plan_name and opens it."""
        export_dir = get_permanent_directory("Export Summaries")
        os.makedirs(export_dir, exist_ok=True)

        # Access the loading_plan_name
        loading_plan_name = getattr(self.parent.parent, 'loading_plan_name', None)

        # Create a filename with timestamp, and include loading_plan_name if available
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        if loading_plan_name:
            sanitized_name = self.sanitize_filename(loading_plan_name)
            filename = os.path.join(export_dir, f"Loading_Summary_{sanitized_name}.xlsx")
        else:
            filename = os.path.join(export_dir, f"Loading_Summary_{timestamp}.xlsx")

        try:
            wb.save(filename)
            
            # Open the file only if save was successful
            self.open_file(filename)
            
            container_ids = sorted(self.parent.containers_packed_items.keys())
            if container_ids:
                first_container_id = container_ids[0]
                self.parent.display_container(first_container_id)
                self.parent.current_container_index = 0
                self.parent.left_panel.container_label.setText(f"Container 1 of {self.parent.total_containers}")
                self.parent.update_navigation_buttons()

            QMessageBox.information(self.parent, "Export Successful", f"Loading summary exported to {filename}")
            
        except PermissionError:
            # Specific handler for permission errors (file already open)
            QMessageBox.warning(
                self.parent, 
                "File Access Error", 
                f"Could not save to {filename}.\n\nThe file may be open in Excel or another program. "
                "Please close the file and try again."
            )
        except Exception as e:
            # General error handler
            QMessageBox.critical(
                self.parent,
                "Export Failed",
                f"An error occurred while saving the workbook:\n{str(e)}"
            )

    # ------------- Metric Calculations -------------

    def calculate_volume_use_rate(self, container, packed_items):
        if not container:
            return "0.00%"
        container_volume = container.length * container.width * container.height
        used_volume = sum(pi.size[0] * pi.size[1] * pi.size[2] for pi in packed_items)
        rate = (used_volume / container_volume * 100) if container_volume > 0 else 0
        return f"{rate:.2f}%"

    def calculate_weight_use_rate(self, container, packed_items):
        if not container or container.max_weight <= 0:
            return "0.00%"
        used_weight = sum(pi.weight for pi in packed_items)
        rate = (used_weight / container.max_weight) * 100
        return f"{rate:.2f}%"

    def calculate_goods_quantity(self, aggregated):
        return sum(d['quantity'] for d in aggregated.values())

    def calculate_goods_volume(self, aggregated):
        total_cm3 = sum(d['length'] * d['width'] * d['height'] * d['quantity'] for d in aggregated.values())
        return f"{total_cm3 / 1_000_000:.2f} m³"

    def calculate_goods_weight(self, aggregated):
        return sum(d['weight'] for d in aggregated.values())

    def calculate_remainder_lengthwise(self, container, packed_items):
        if not container:
            return "N/A"
        if not packed_items:
            return container.length
        max_extent = max(pi.position[0] + pi.size[0] for pi in packed_items)
        remainder = container.length - max_extent
        return max(remainder, 0.0)

    def calculate_remainder_widthwise(self, container, packed_items):
        if not container:
            return "N/A"
        if not packed_items:
            return container.width
        max_extent = max(pi.position[1] + pi.size[1] for pi in packed_items)
        remainder = container.width - max_extent
        return max(remainder, 0.0)

    def calculate_remainder_heightwise(self, container, packed_items):
        if not container:
            return "N/A"
        if not packed_items:
            return container.height
        max_extent = max(pi.position[2] + pi.size[2] for pi in packed_items)
        remainder = container.height - max_extent
        return max(remainder, 0.0)

    # ------------- Error & File Handling -------------

    def handle_exception(self, e):
        QMessageBox.critical(self.parent, "Export Failed", f"An error occurred while exporting:\n{e}")

    def open_file(self, file_path):
        """Opens the given file in the default application."""
        try:
            if sys.platform.startswith('darwin'):
                subprocess.call(['open', file_path])
            elif os.name == 'nt':
                os.startfile(file_path)
            elif os.name == 'posix':
                subprocess.call(['xdg-open', file_path])
        except Exception as ex:
            QMessageBox.warning(self.parent, "Open File Failed", f"Could not open the file:\n{ex}")

    def sanitize_filename(self, name):
        """Removes invalid characters from a filename."""
        return re.sub(r'[^a-zA-Z0-9_\-]', '_', name)

    # ----------------------------
    # 2) Istia Export (JSON)
    # ----------------------------

    def create_istia_export(self):
        """
        Exports container data, item data, loading order, and per-container metrics
        into a JSON file for external integration (e.g., Istia).
        """
        try:
            # Collect all necessary data using existing methods
            data = self.collect_export_data_for_istia()
            
            # Define the export directory and ensure it exists
            export_dir = get_permanent_directory("Istia Exports")
            os.makedirs(export_dir, exist_ok=True)
    
            # Access the loading_plan_name from the appropriate parent
            loading_plan_name = getattr(self.parent.parent, 'loading_plan_name', None)
    
            # Create a filename with timestamp, and include loading_plan_name if available
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            if loading_plan_name:
                sanitized_name = self.sanitize_filename(loading_plan_name)
                filename = os.path.join(export_dir, f"Istia_Export_{sanitized_name}.json")
            else:
                filename = os.path.join(export_dir, f"Istia_Export_{timestamp}.json")
    
            # Write the collected data to the JSON file
            with open(filename, 'w') as file:
                json.dump(data, file, indent=4)
    
            # Inform the user of a successful export
            QMessageBox.information(self.parent, "Export Successful", f"Istia export saved to:\n{filename}")
    
            # Open the export directory for the user
            open_folder(export_dir)

        except Exception as e:
            # Handle any exceptions that occur during the export process
            self.handle_exception(e)

    def collect_export_data_for_istia(self):
        """
        Gathers all essential data for an ISTIA-like JSON export,
        including container data, packed items, and separate lists
        for all combined, euro, and mixed pallets across all containers.
        Pulls contained_items from data_manager to ensure
        we capture the actual nested items for combined/mixed pallets.
        For contained items, only the minimal fields are included.
        """
        from datetime import datetime
        data_manager = self.parent.parent.data_manager
        plan_name = getattr(self.parent.parent.parent, 'loading_plan_name', None) or "UnnamedPlan"

        # Helper function: return only minimal fields for a contained item
        def minimal_item_dict(item):
            return {
                "sku": item.sku,
                "length": item.length,
                "width": item.width,
                "height": item.height,
                "weight": item.weight,
                "quantity": item.quantity
            }

        # Build a dictionary of original items for easy lookup by SKU
        items_map = {it.sku: it for it in data_manager.items}

        # Prepare root export structure
        data = {
            "loading_plan_name": plan_name,
            "timestamp": datetime.now().isoformat(),
            "margin_percentage": data_manager.margin_percentage,
            "items": [],            # Full list of items from items_map
            "combined_pallets": [], # Overviews for all combined pallets across containers
            "combined_europallets": [],      # Overviews for all combined europallets across containers
            "mixed_pallets": [],    # Overviews for all mixed pallets across containers
            "containers": []
        }

        # 1) Add all items from items_map to the global "items" section
        for sku, item in items_map.items():
            data["items"].append({
                "sku": sku,
                "size_cm": {
                    "length": item.length,
                    "width": item.width,
                    "height": item.height
                },
                "weight_kg": item.weight,
                "quantity": item.quantity,
                "stackable": item.stackable,
                "rotatable": item.rotatable,
                "europallet": item.europallet,
                "mixed_pallet": item.mixed_pallet,
                "cartons": item.cartons,
                "transport_order": item.transport_order,
                # Exclude: "items_per_pallet", "has_carton_issue", "has_remainder_issue", "has_missing_dimension_issue", "contained_items", 
                #          "original_quantity", "is_carton_item"
                "contained_items": [ci.to_dict() for ci in item.contained_items]
            })

        # 2) Gather container + packed_items data from the VisualizationPage
        containers_map = self.parent.containers           # {container_id: Container}
        packed_map = self.parent.containers_packed_items  # {container_id: [PackedItem, ...]}

        # Track which synthetic pallets we've already added to the top-level arrays
        processed_combined_pallets = set()
        processed_europallets = set()
        processed_mixed_pallets = set()

        # 3) Populate container information and link packed items
        for cid in sorted(packed_map.keys()):
            container = containers_map.get(cid)
            if container is None:
                continue  # Skip if no container details

            # Compute container metrics
            container_packed_items = packed_map[cid]
            aggregated = self.aggregate_packed_items(container_packed_items)
            metrics_list = self.calculate_metrics(container, container_packed_items, aggregated)
            metrics_dict = {label: value for (label, value) in metrics_list}

            # Build container dict
            container_data = {
                "container_id": cid,
                "container_type": container.container_type,
                "length_cm": container.length,
                "width_cm": container.width,
                "height_cm": container.height,
                "max_weight_kg": container.max_weight,
                "metrics": metrics_dict,
                "packed_items": []
            }

            # Sort items by (x,y,z) for stable loading order
            sorted_packed_items = sorted(
                container_packed_items,
                key=lambda pi: (pi.position[0], pi.position[1], pi.position[2])
            )

            # 4) Add each packed item to the container's "packed_items" list
            for idx_item, p_item in enumerate(sorted_packed_items, start=1):
                item_dict = {
                    "order": idx_item,
                    "sku": p_item.sku,
                    "position_cm": {
                        "x": p_item.position[0],
                        "y": p_item.position[1],
                        "z": p_item.position[2]
                    },
                    "size_cm": {
                        "length": p_item.size[0],
                        "width": p_item.size[1],
                        "height": p_item.size[2]
                    },
                    "weight_kg": p_item.weight,
                    "rotation_degrees": p_item.rotation,
                }

                if p_item.sku.startswith("CombP-") or p_item.sku.startswith("EuroP-") or p_item.sku.startswith("MIXED-"):
                    item_dict["is_synthetic_pallet"] = True
                    # 4a) Attempt to look up the "real" item from data_manager.items or combined_pallets
                    real_item = items_map.get(p_item.sku, None)
                    if real_item is None:
                        # fallback: look in data_manager.combined_pallets
                        real_item = next(
                            (cp for cp in data_manager.combined_pallets if cp.sku == p_item.sku),
                            None
                        )
                    if real_item is not None:
                        # Use minimal_item_dict for each contained item
                        item_dict["contained_items"] = [minimal_item_dict(ci) for ci in real_item.contained_items]
                    else:
                        item_dict["contained_items"] = []
                    
                    # 4b) Insert into top-level combined_pallets, europallets or mixed_pallets if not processed yet
                    if p_item.sku.startswith("CombP-") and p_item.sku not in processed_combined_pallets:
                        data["combined_pallets"].append({
                            "sku": p_item.sku,
                            "weight_kg": p_item.weight,
                            "is_synthetic_pallet": True,
                            "contained_items": item_dict["contained_items"]
                        })
                        processed_combined_pallets.add(p_item.sku)
                    elif p_item.sku.startswith("EuroP-") and p_item.sku not in processed_europallets:
                        data["combined_europallets"].append({
                            "sku": p_item.sku,
                            "weight_kg": p_item.weight,
                            "is_synthetic_pallet": True,
                            "contained_items": item_dict["contained_items"],
                            "pallet_type": "euro"
                        })
                        processed_europallets.add(p_item.sku)
                    elif p_item.sku.startswith("MIXED-") and p_item.sku not in processed_mixed_pallets:
                        data["mixed_pallets"].append({
                            "sku": p_item.sku,
                            "weight_kg": p_item.weight,
                            "is_synthetic_pallet": True,
                            "contained_items": item_dict["contained_items"]
                        })
                        processed_mixed_pallets.add(p_item.sku)
        
                else:
                    # Normal item
                    item_dict["is_synthetic_pallet"] = False
                    item_dict["sku_reference"] = p_item.sku
                    # Normal items do not have contained_items

                container_data["packed_items"].append(item_dict)

            # Finally, append container_data
            data["containers"].append(container_data)

        return data


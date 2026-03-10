# test_summary_exporter.py

import pytest
from unittest.mock import MagicMock, patch, mock_open
from summary_exporter import SummaryExporter
import openpyxl
import os

@pytest.fixture
def mock_parent():
    """
    Creates a mock parent object with necessary attributes and methods.
    """
    parent = MagicMock()

    # Mock the containers_packed_items with sample data
    parent.containers_packed_items = {
        'C1': [
            MagicMock(sku='SKU1', position=(10, 20, 30), size=(100, 200, 300), weight=10.5),
            MagicMock(sku='SKU2', position=(40, 50, 60), size=(150, 250, 350), weight=20.5),
        ],
        'C2': [
            MagicMock(sku='SKU1', position=(15, 25, 35), size=(110, 210, 310), weight=11.5),
        ]
    }

    # Mock the containers with container_type and dimensions
    parent.containers = {
        'C1': MagicMock(container_type='TypeA', length=1000, width=1000, height=1000, max_weight=500),
        'C2': MagicMock(container_type='TypeB', length=1200, width=1100, height=1300, max_weight=600),
    }

    parent.total_containers = 2
    parent.current_container_index = 0

    # Mock methods related to display and UI updates
    parent.display_container = MagicMock()
    parent.right_panel.view.grabFramebuffer.return_value = MagicMock(save=MagicMock())
    parent.right_panel.view.setCameraPosition = MagicMock()
    parent.right_panel.update_wall_visibility = MagicMock()
    parent.left_panel.container_label = MagicMock()
    parent.left_panel.container_label.setText = MagicMock()
    parent.update_navigation_buttons = MagicMock()

    # Mock loading_plan_name attribute
    parent.parent.loading_plan_name = "TestPlan"

    return parent

@pytest.fixture
def exporter(mock_parent):
    """
    Initializes the SummaryExporter with the mocked parent.
    """
    return SummaryExporter(parent=mock_parent)

@patch('summary_exporter.get_permanent_directory')
@patch('summary_exporter.QMessageBox')
@patch('summary_exporter.openpyxl.Workbook')
@patch('summary_exporter.sys')
@patch('summary_exporter.os')
def test_export_loading_summary_success(mock_os, mock_sys, mock_Workbook, mock_QMessageBox, mock_get_permanent_directory, exporter, mock_parent):
    """
    Tests the export_loading_summary method for successful export.
    """
    # Setup mocks
    mock_wb = MagicMock()
    mock_ws = MagicMock()
    mock_loading_order_ws = MagicMock()
    mock_Workbook.return_value = mock_wb
    mock_wb.active = mock_ws
    mock_wb.create_sheet.return_value = mock_loading_order_ws

    mock_get_permanent_directory.return_value = "/fake/export/path"

    # Mock the save method
    mock_wb.save = MagicMock()

    # Execute the method
    exporter.export_loading_summary()

    # Assertions
    mock_Workbook.assert_called_once()
    mock_wb.create_sheet.assert_called_with(title="Loading Order")
    mock_ws.title = "Loading Summary"
    mock_ws.sheet_view.showGridLines = False
    mock_loading_order_ws.sheet_view.showGridLines = False

    # Check if save was called with the correct file path
    expected_file_path = os.path.join("/fake/export/path", "Loading_Summary_TestPlan.xlsx")
    mock_wb.save.assert_called_with(expected_file_path)

    # Check if QMessageBox.information was called for successful export
    mock_QMessageBox.information.assert_called_with(mock_parent, "Export Successful", f"Loading summary exported to {expected_file_path}")

    # Check if the first container is displayed after export
    mock_parent.display_container.assert_called_with('C1')
    mock_parent.left_panel.container_label.setText.assert_called_with("Container 1 of 2")
    mock_parent.update_navigation_buttons.assert_called_once()

@patch('summary_exporter.get_permanent_directory')
@patch('summary_exporter.QMessageBox')
def test_export_loading_summary_no_data(mock_QMessageBox, mock_get_permanent_directory, exporter, mock_parent):
    """
    Tests the export_loading_summary method when there is no data to export.
    """
    # Setup mocks
    mock_parent.containers_packed_items = {}

    # Execute the method
    exporter.export_loading_summary()

    # Assertions
    mock_QMessageBox.information.assert_called_with(mock_parent, "No Data", "No containers to export.")

@patch('summary_exporter.get_permanent_directory')
@patch('summary_exporter.QMessageBox')
@patch('summary_exporter.openpyxl.Workbook')
def test_export_loading_summary_exception(mock_Workbook, mock_QMessageBox, mock_get_permanent_directory, exporter, mock_parent):
    """
    Tests the export_loading_summary method when an exception occurs.
    """
    # Setup mocks to raise an exception
    mock_Workbook.side_effect = Exception("Workbook creation failed")

    # Execute the method
    exporter.export_loading_summary()

    # Assertions
    mock_QMessageBox.critical.assert_called_with(mock_parent, "Export Failed", "An error occurred while exporting:\nWorkbook creation failed")

@patch('summary_exporter.get_permanent_directory')
@patch('summary_exporter.QMessageBox')
@patch('summary_exporter.openpyxl.Workbook')
@patch('summary_exporter.os.startfile')
def test_finalize_workbook_os_specific(mock_startfile, mock_Workbook, mock_QMessageBox, mock_get_permanent_directory, exporter, mock_parent):
    """
    Tests the finalize_workbook method for different OS platforms.
    """
    # Setup
    mock_wb = MagicMock()
    mock_ws = MagicMock()
    mock_loading_order_ws = MagicMock()
    mock_Workbook.return_value = mock_wb
    mock_wb.active = mock_ws
    mock_wb.create_sheet.return_value = mock_loading_order_ws

    mock_get_permanent_directory.return_value = "/fake/export/path"
    mock_parent.parent.loading_plan_name = "FinalPlan"

    file_path = os.path.join("/fake/export/path", "Loading_Summary_FinalPlan.xlsx")

    # Call finalize_workbook directly
    with patch.object(exporter, 'open_file') as mock_open_file:
        exporter.finalize_workbook(mock_wb, mock_ws, mock_loading_order_ws)

        # Assertions
        mock_wb.save.assert_called_with(file_path)
        mock_open_file.assert_called_with(file_path)
        mock_QMessageBox.information.assert_called_with(mock_parent, "Export Successful", f"Loading summary exported to {file_path}")

@patch('summary_exporter.get_permanent_directory')
@patch('summary_exporter.os.startfile')
def test_open_file_windows(mock_startfile, mock_get_permanent_directory, exporter, mock_parent):
    """
    Tests the open_file method on Windows platform.
    """
    mock_get_permanent_directory.return_value = "/fake/export/path"

    # Set sys.platform to Windows
    with patch('summary_exporter.sys.platform', 'win32'):
        with patch.object(exporter, 'open_file') as mock_open_file_method:
            exporter.open_file("C:\\path\\to\\file.xlsx")
            mock_open_file_method.assert_not_called()  # Because we are mocking open_file itself

    # Alternatively, to test internal behavior, we should not patch open_file
    with patch('summary_exporter.sys.platform', 'win32'):
        exporter.open_file("C:\\path\\to\\file.xlsx")
        mock_startfile.assert_called_with("C:\\path\\to\\file.xlsx")

@patch('summary_exporter.get_permanent_directory')
@patch('summary_exporter.os.startfile')
def test_open_file_non_windows(mock_startfile, mock_get_permanent_directory, exporter, mock_parent):
    """
    Tests the open_file method on non-Windows platforms.
    """
    mock_get_permanent_directory.return_value = "/fake/export/path"

    # Test for macOS
    with patch('summary_exporter.sys.platform', 'darwin'):
        with patch('summary_exporter.subprocess.call') as mock_subprocess_call:
            exporter.open_file("/path/to/file.xlsx")
            mock_subprocess_call.assert_called_with(['open', '/path/to/file.xlsx'])

    # Test for Linux
    with patch('summary_exporter.sys.platform', 'linux'):
        with patch('summary_exporter.subprocess.call') as mock_subprocess_call:
            exporter.open_file("/path/to/file.xlsx")
            mock_subprocess_call.assert_called_with(['xdg-open', '/path/to/file.xlsx'])

@patch('summary_exporter.get_permanent_directory')
@patch('summary_exporter.QMessageBox')
def test_handle_exception(mock_QMessageBox, mock_get_permanent_directory, exporter, mock_parent):
    """
    Tests the handle_exception method.
    """
    # Create an exception
    exception = Exception("Test exception")

    # Call handle_exception
    exporter.handle_exception(exception)

    # Assertions
    mock_QMessageBox.critical.assert_called_with(mock_parent, "Export Failed", f"An error occurred while exporting:\n{exception}")

def test_calculate_metrics(exporter, mock_parent):
    """
    Tests the calculate_metrics method.
    """
    container = mock_parent.containers['C1']
    packed_items = mock_parent.containers_packed_items['C1']
    aggregated_items = exporter.aggregate_packed_items(packed_items)

    metrics = exporter.calculate_metrics(container, packed_items, aggregated_items)

    expected_metrics = [
        ("Volume Use Rate (%)", "630.00%"),  # (1000*1000*1000) vs sum of packed volumes
        ("Weight Use Rate (%)", "6.20%"),   # (10.5 + 20.5 + 11.5) / 500 * 100
        ("Goods Quantity", 3),
        ("Goods Volume (m³)", "0.06 m³"),   # sum of volumes / 1_000_000
        ("Goods Weight (kg)", 42.5),
        ("Remainder Lengthwise", 0.0),
        ("Remainder Widthwise", 0.0),
        ("Remainder Heightwise", 0.0),
    ]

    assert metrics == expected_metrics
    
